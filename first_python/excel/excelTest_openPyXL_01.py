from openpyxl import Workbook

wb = Workbook()

ws = wb.active
ws = wb.create_sheet("diary", 0) # 0이면 첫번째 위치에 삽입, 생략시 마지막 위치에 삽입

data = [('김초롱', 88, 99, 100), ('최치원', 80, 70, 90), ('홍길동', 80, 70, 90)]

r =1 # 첫 번째 행이 1부터 시작
c =1 # 첫 번째 열이 1부터 시작

for irum, kor, eng, math in data:
    ws.cell(row=r, column=c).value = irum   # ws.cell(row=r, column=c, value=irum)
    ws.cell(row=r, column=c+1).value = kor
    ws.cell(row=r, column=c+2).value = eng
    ws.cell(row=r, column=c+3).value = math
    r += 1

ws.cell(row=r, column=1).value = '합계'
ws.cell(row=r, column=2).value = '=sum(B1:B3)'
ws.cell(row=r, column=3).value = '=sum(C1:C3)'
ws.cell(row=r, column=4).value = '=sum(D1:D3)'

wb.save("openpyxl22.xlsx")
wb.close()
