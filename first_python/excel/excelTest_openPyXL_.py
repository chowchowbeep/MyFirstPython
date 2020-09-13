# https://openpyxl.readthedocs.io/en/stable/ 사이트참조
# 설치(터미널) : pip install openpyxl
# 설치(gui) : file - settings - project - python interpreter - +버튼 - 패키지 검색 및 선택 - install
from openpyxl import Workbook, load_workbook

wb = Workbook() # 엑셀 문서 객체를 생성함
ws = wb.active # 활성화되어 있는 시트 선택
# ws wb.create_sheet(0) # 시트생성 # 0이면 첫번째 위치에 삽입, 생략시 마지막 위치에 삽입
ws.title = "sample" # 워크시트 타이틀 설정
ws['B2'] = 42 # B2 cell에 데이터 삽입
ws.append([1, 2, 3]) # 마지막 위치 아래 행에 데이터 삽입
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([4, 4, 4])
wb.save("openpyxl1.xlsx") # 파일저장
wb.close() # 파일닫기


wb = load_workbook(filename ='openpyxl1.xlsx') # 파일 열기
ws = wb.active
ws['A1'] = 42
ws.cell(row=1, column=3).value = 333 # row, column으로 셀 지정하여 데이터 삽입
ws.append(['aaa', 'bbb', 'ccc'])

print(ws['A1'].value) # A1셀의 데이터 가져오기 # 저장된 값이 없으면 None으로 출력
print(ws['A2'].value)

ws2 = wb['sample'] # sample 시트 선택
print(ws2['A3'].value, ws2['B3'].value, ws2['C3'].value)
print(ws2['A4'].value, ws2['B4'].value, ws2['C4'].value)
print(ws2['A5'].value, ws2['B5'].value, ws2['C5'].value)

wb.save("openpyxl1.xlsx")